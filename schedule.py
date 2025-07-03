import os
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl is not installed. Please install it to use this script.")
    sys.exit(1)

FILE_NAME = "horario.xlsx"

def main():
    if os.path.exists(FILE_NAME):
        wb = load_workbook(FILE_NAME)
        print(f"{FILE_NAME} found. Sheet names: {wb.sheetnames}")
    else:
        wb = Workbook()
        wb.save(FILE_NAME)
        print(f"{FILE_NAME} created.")

if __name__ == "__main__":
    main()
